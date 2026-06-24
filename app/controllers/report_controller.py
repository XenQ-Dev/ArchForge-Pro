"""
Report Controller — professional PDF reports with embedded charts.
Charts generated with matplotlib and embedded as PNG images via ReportLab.
"""
from __future__ import annotations
import io
from pathlib import Path
from datetime import datetime

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.gridspec import GridSpec
import numpy as np

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable, Image, KeepTogether,
)
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.lib.utils import ImageReader

from app.models.project_model import get_project
from app.models.estimate_model import get_estimate
from app.models.expense_model import get_expenses, get_expense_summary, get_monthly_spending
from app.models.boq_model import get_boq
from app.models.timeline_model import get_phases
from app.models.settings_model import get_all_settings
from app.utils.formatters import fmt_inr, fmt_date

# ── Palette ──────────────────────────────────────────────────────────────────
C_NAVY    = colors.HexColor("#1e3a5f")
C_GOLD    = colors.HexColor("#c8973a")
C_STEEL   = colors.HexColor("#2c4a6e")
C_LIGHT   = colors.HexColor("#f0f4f8")
C_WHITE   = colors.white
C_DARK    = colors.HexColor("#1a1d23")
C_GREEN   = colors.HexColor("#2e7d4f")
C_RED     = colors.HexColor("#b83232")
C_ORANGE  = colors.HexColor("#c8620a")

# Matplotlib chart colours
_CHART_PALETTE = ["#1e3a5f", "#c8973a", "#2c7873", "#c8620a", "#6c4a8a", "#2e7d4f", "#b83232"]

PAGE_W, PAGE_H = A4
MARGIN = 15 * mm
USABLE_W = PAGE_W - 2 * MARGIN


# ── Document helpers ─────────────────────────────────────────────────────────
def _doc(path: str, title: str) -> SimpleDocTemplate:
    return SimpleDocTemplate(
        path, pagesize=A4,
        rightMargin=MARGIN, leftMargin=MARGIN,
        topMargin=16 * mm, bottomMargin=16 * mm,
        title=title, author="ArchForge Pro",
    )


def _styles():
    s = getSampleStyleSheet()
    add = s.add
    add(ParagraphStyle("RF_Title",    fontName="Helvetica-Bold", fontSize=20,
                       textColor=C_NAVY, alignment=TA_CENTER, spaceAfter=2))
    add(ParagraphStyle("RF_Sub",      fontName="Helvetica", fontSize=10,
                       textColor=C_GOLD, alignment=TA_CENTER, spaceAfter=2))
    add(ParagraphStyle("RF_Address",  fontName="Helvetica", fontSize=8,
                       textColor=C_DARK, alignment=TA_CENTER, spaceAfter=4))
    add(ParagraphStyle("RF_Section",  fontName="Helvetica-Bold", fontSize=11,
                       textColor=C_NAVY, spaceBefore=12, spaceAfter=4,
                       borderPad=2))
    add(ParagraphStyle("RF_Body",     fontName="Helvetica", fontSize=9,
                       textColor=C_DARK, leading=14))
    add(ParagraphStyle("RF_Note",     fontName="Helvetica-Oblique", fontSize=8,
                       textColor=colors.HexColor("#555555"), leading=12))
    add(ParagraphStyle("RF_Right",    fontName="Helvetica", fontSize=9,
                       textColor=C_DARK, alignment=TA_RIGHT))
    add(ParagraphStyle("RF_Caption",  fontName="Helvetica-Oblique", fontSize=8,
                       textColor=colors.HexColor("#555555"), alignment=TA_CENTER,
                       spaceBefore=2, spaceAfter=6))
    return s


def _header(settings: dict, s, report_name: str, project: dict) -> list:
    company = settings.get("company_name", "ArchForge Pro")
    address = settings.get("company_address", "")
    phone   = settings.get("company_phone", "")
    gst     = settings.get("gst_number", "")
    contact_line = "  |  ".join(x for x in [phone, gst] if x)
    elements = [
        Paragraph(company.upper(), s["RF_Title"]),
        Paragraph("Construction Cost Estimation & Project Management", s["RF_Sub"]),
    ]
    if address:
        elements.append(Paragraph(address, s["RF_Address"]))
    if contact_line:
        elements.append(Paragraph(contact_line, s["RF_Address"]))
    elements += [
        HRFlowable(width="100%", thickness=2, color=C_NAVY, spaceAfter=2),
        HRFlowable(width="100%", thickness=1, color=C_GOLD, spaceAfter=8),
        Paragraph(report_name, s["RF_Section"]),
        Paragraph(
            f"Generated on {datetime.now().strftime('%d %B %Y, %I:%M %p')}  "
            f"  |  Project: <b>{project['project_name']}</b>  "
            f"  |  Client: <b>{project['client_name']}</b>",
            s["RF_Body"]
        ),
        Spacer(1, 6),
    ]
    return elements


def _project_block(project: dict, s) -> KeepTogether:
    data = [
        [Paragraph("<b>Project Name</b>", s["RF_Body"]),  project["project_name"],
         Paragraph("<b>Client</b>", s["RF_Body"]),         project["client_name"]],
        [Paragraph("<b>Type</b>", s["RF_Body"]),           project["project_type"],
         Paragraph("<b>Quality</b>", s["RF_Body"]),        project["construction_quality"]],
        [Paragraph("<b>Location</b>", s["RF_Body"]),       project.get("site_location","—"),
         Paragraph("<b>Floors</b>", s["RF_Body"]),         str(project["num_floors"])],
        [Paragraph("<b>Built-up Area</b>", s["RF_Body"]),  f"{project['built_up_area']:,.0f} sq.ft",
         Paragraph("<b>Start Date</b>", s["RF_Body"]),     fmt_date(project["start_date"])],
        [Paragraph("<b>Status</b>", s["RF_Body"]),         project["status"],
         Paragraph("<b>Completion</b>", s["RF_Body"]),     fmt_date(project.get("expected_completion","—"))],
    ]
    cw = [38*mm, 56*mm, 38*mm, 56*mm]
    tbl = Table(data, colWidths=cw, repeatRows=0)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), colors.white),
        ("BACKGROUND", (0,0), (0,-1), C_LIGHT),
        ("BACKGROUND", (2,0), (2,-1), C_LIGHT),
        ("FONTSIZE",   (0,0), (-1,-1), 9),
        ("BOX",        (0,0), (-1,-1), 0.8, C_NAVY),
        ("INNERGRID",  (0,0), (-1,-1), 0.3, colors.HexColor("#cccccc")),
        ("ROWBACKGROUNDS", (0,0), (-1,-1), [colors.white, colors.HexColor("#f8f9fb")]),
        ("TOPPADDING",  (0,0), (-1,-1), 5),
        ("BOTTOMPADDING",(0,0),(-1,-1), 5),
        ("LEFTPADDING", (0,0), (-1,-1), 7),
    ]))
    return KeepTogether([Paragraph("Project Details", s["RF_Section"]), tbl, Spacer(1,8)])


def _section_bar(text: str, s) -> list:
    return [
        HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#cccccc"), spaceBefore=4),
        Paragraph(text, s["RF_Section"]),
    ]


def _std_table(headers: list, rows: list, col_widths: list,
               right_cols: list[int] | None = None) -> Table:
    all_rows = [headers] + rows
    style = [
        ("BACKGROUND",  (0,0), (-1,0), C_NAVY),
        ("TEXTCOLOR",   (0,0), (-1,0), C_WHITE),
        ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",    (0,0), (-1,-1), 9),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, C_LIGHT]),
        ("BOX",         (0,0), (-1,-1), 0.5, C_NAVY),
        ("INNERGRID",   (0,0), (-1,-1), 0.3, colors.HexColor("#cccccc")),
        ("TOPPADDING",  (0,0), (-1,-1), 5),
        ("BOTTOMPADDING",(0,0),(-1,-1), 5),
        ("LEFTPADDING", (0,0), (-1,-1), 6),
        ("ALIGN",       (0,0), (-1,0), "CENTER"),
    ]
    for c in (right_cols or []):
        style.append(("ALIGN", (c,1), (c,-1), "RIGHT"))
    tbl = Table(all_rows, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle(style))
    return tbl


def _total_row(label: str, value: str, width: float = USABLE_W) -> Table:
    left_w = width - 45*mm
    tbl = Table([[label, value]], colWidths=[left_w, 45*mm])
    tbl.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,-1), C_NAVY),
        ("TEXTCOLOR",    (0,0), (-1,-1), C_WHITE),
        ("FONTNAME",     (0,0), (-1,-1), "Helvetica-Bold"),
        ("FONTSIZE",     (0,0), (-1,-1), 11),
        ("ALIGN",        (1,0), (1,0), "RIGHT"),
        ("TOPPADDING",   (0,0), (-1,-1), 8),
        ("BOTTOMPADDING",(0,0), (-1,-1), 8),
        ("LEFTPADDING",  (0,0), (-1,-1), 8),
    ]))
    return tbl


def _highlight_row(label: str, value: str, col: colors.Color = C_GOLD) -> Table:
    tbl = Table([[label, value]], colWidths=[USABLE_W - 45*mm, 45*mm])
    tbl.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,-1), col),
        ("TEXTCOLOR",    (0,0), (-1,-1), C_DARK),
        ("FONTNAME",     (0,0), (-1,-1), "Helvetica-Bold"),
        ("FONTSIZE",     (0,0), (-1,-1), 10),
        ("ALIGN",        (1,0), (1,0), "RIGHT"),
        ("TOPPADDING",   (0,0), (-1,-1), 6),
        ("BOTTOMPADDING",(0,0), (-1,-1), 6),
        ("LEFTPADDING",  (0,0), (-1,-1), 8),
    ]))
    return tbl


# ── Chart helpers ─────────────────────────────────────────────────────────────
def _chart_to_image(fig, width_mm: float, height_mm: float) -> Image:
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight",
                facecolor="white", edgecolor="none")
    plt.close(fig)
    buf.seek(0)
    return Image(buf, width=width_mm * mm, height=height_mm * mm)


def _pie_chart(labels: list, values: list, title: str,
               w_mm=85, h_mm=70) -> Image:
    fig, ax = plt.subplots(figsize=(w_mm/25.4, h_mm/25.4))
    wedges, texts, autotexts = ax.pie(
        values, labels=None, colors=_CHART_PALETTE[:len(values)],
        autopct="%1.1f%%", startangle=140, pctdistance=0.78,
        wedgeprops=dict(linewidth=1.2, edgecolor="white"),
    )
    for t in autotexts:
        t.set_fontsize(7)
        t.set_color("white")
        t.set_fontweight("bold")
    ax.legend(wedges, labels, loc="lower center", bbox_to_anchor=(0.5, -0.18),
              ncol=2, fontsize=7, frameon=False)
    ax.set_title(title, fontsize=9, fontweight="bold", color="#1e3a5f", pad=8)
    fig.tight_layout()
    return _chart_to_image(fig, w_mm, h_mm)


def _bar_chart(categories: list, values: list, title: str, ylabel: str = "₹",
               w_mm=170, h_mm=70, color_list=None) -> Image:
    fig, ax = plt.subplots(figsize=(w_mm/25.4, h_mm/25.4))
    cols = color_list or _CHART_PALETTE[:len(categories)]
    bars = ax.bar(categories, values, color=cols, edgecolor="white",
                  linewidth=0.8, zorder=3)
    ax.yaxis.grid(True, linestyle="--", alpha=0.5, zorder=0)
    ax.set_axisbelow(True)
    ax.spines[["top","right"]].set_visible(False)
    ax.spines[["left","bottom"]].set_color("#cccccc")
    ax.tick_params(axis="x", labelsize=7, rotation=15)
    ax.tick_params(axis="y", labelsize=7)
    ax.set_ylabel(ylabel, fontsize=8, color="#555555")
    ax.set_title(title, fontsize=9, fontweight="bold", color="#1e3a5f", pad=8)
    for bar, val in zip(bars, values):
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() * 1.01,
                f"₹{val/1e5:.1f}L" if val >= 1e5 else f"₹{val:,.0f}",
                ha="center", va="bottom", fontsize=6.5, color="#333333")
    fig.tight_layout()
    return _chart_to_image(fig, w_mm, h_mm)


def _line_chart(x_labels: list, values: list, title: str,
                w_mm=170, h_mm=65) -> Image:
    fig, ax = plt.subplots(figsize=(w_mm/25.4, h_mm/25.4))
    ax.plot(x_labels, values, color=_CHART_PALETTE[0], linewidth=2,
            marker="o", markersize=5, markerfacecolor=_CHART_PALETTE[1],
            markeredgecolor="white", markeredgewidth=1)
    ax.fill_between(x_labels, values, alpha=0.12, color=_CHART_PALETTE[0])
    ax.yaxis.grid(True, linestyle="--", alpha=0.4)
    ax.set_axisbelow(True)
    ax.spines[["top","right"]].set_visible(False)
    ax.spines[["left","bottom"]].set_color("#cccccc")
    ax.tick_params(axis="x", labelsize=7, rotation=30)
    ax.tick_params(axis="y", labelsize=7)
    ax.set_title(title, fontsize=9, fontweight="bold", color="#1e3a5f", pad=8)
    fig.tight_layout()
    return _chart_to_image(fig, w_mm, h_mm)


def _gantt_chart(phases: list, w_mm=170, h_mm=90) -> Image:
    """Horizontal Gantt bar chart for project phases."""
    names   = [p["phase_name"] for p in phases]
    planned = []
    actual  = []
    pct     = [p.get("completion_pct", 0) or 0 for p in phases]

    def _parse(d):
        if not d:
            return None
        try:
            return datetime.strptime(str(d), "%Y-%m-%d")
        except Exception:
            return None

    ref = None
    for p in phases:
        s = _parse(p.get("planned_start"))
        if s and (ref is None or s < ref):
            ref = s
    if ref is None:
        ref = datetime.today()

    def _days(d):
        dt = _parse(d)
        return (dt - ref).days if dt else None

    for p in phases:
        ps, pe = _days(p.get("planned_start")), _days(p.get("planned_end"))
        as_, ae = _days(p.get("actual_start")),  _days(p.get("actual_end"))
        planned.append((ps, pe))
        actual.append((as_, ae))

    fig, ax = plt.subplots(figsize=(w_mm/25.4, h_mm/25.4))
    y_pos = list(range(len(names)))

    for i, (ps, pe) in enumerate(planned):
        if ps is not None and pe is not None:
            ax.barh(i, pe - ps, left=ps, height=0.35, color="#d0d8e8",
                    edgecolor="#1e3a5f", linewidth=0.6, zorder=2)

    for i, (as_, ae) in enumerate(actual):
        if as_ is not None and ae is not None:
            ax.barh(i, ae - as_, left=as_, height=0.35, color=_CHART_PALETTE[0],
                    edgecolor="white", linewidth=0.5, zorder=3, alpha=0.85)
        elif as_ is not None:
            # in progress — show partial bar based on completion
            bar_len = max(5, pct[i] * 0.5)
            ax.barh(i, bar_len, left=as_, height=0.35, color=_CHART_PALETTE[1],
                    edgecolor="white", linewidth=0.5, zorder=3, alpha=0.9)

    ax.set_yticks(y_pos)
    ax.set_yticklabels(names, fontsize=7)
    ax.set_xlabel("Days from project start", fontsize=7, color="#555555")
    ax.set_title("Phase Timeline (Planned vs Actual)", fontsize=9,
                 fontweight="bold", color="#1e3a5f", pad=8)
    ax.spines[["top","right"]].set_visible(False)
    ax.spines[["left","bottom"]].set_color("#cccccc")
    ax.tick_params(axis="x", labelsize=7)
    ax.xaxis.grid(True, linestyle="--", alpha=0.4)
    ax.set_axisbelow(True)
    legend = [
        mpatches.Patch(color="#d0d8e8", edgecolor="#1e3a5f", label="Planned"),
        mpatches.Patch(color=_CHART_PALETTE[0], label="Actual"),
        mpatches.Patch(color=_CHART_PALETTE[1], label="In Progress"),
    ]
    ax.legend(handles=legend, fontsize=6.5, loc="lower right", frameon=False)
    fig.tight_layout()
    return _chart_to_image(fig, w_mm, h_mm)


def _side_by_side(img1: Image, caption1: str, img2: Image, caption2: str, s) -> Table:
    half = USABLE_W / 2 - 4*mm
    img1.drawWidth  = half
    img1.drawHeight = img1.drawHeight * (half / img1.drawWidth) if img1.drawWidth else img1.drawHeight
    img2.drawWidth  = half
    img2.drawHeight = img2.drawHeight * (half / img2.drawWidth) if img2.drawWidth else img2.drawHeight
    tbl = Table(
        [[img1, img2],
         [Paragraph(caption1, s["RF_Caption"]), Paragraph(caption2, s["RF_Caption"])]],
        colWidths=[half + 3*mm, half + 3*mm],
    )
    tbl.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"TOP"),
                              ("LEFTPADDING",(0,0),(-1,-1),2),
                              ("RIGHTPADDING",(0,0),(-1,-1),2)]))
    return tbl


# ── ReportController ─────────────────────────────────────────────────────────
class ReportController:
    def __init__(self):
        self._settings = get_all_settings()

    # ── 1. Cost Estimate Report ──────────────────────────────────────────────
    def generate_cost_estimate_report(self, project_id: int, path: str) -> None:
        project  = get_project(project_id)
        estimate = get_estimate(project_id)
        s        = _styles()
        doc      = _doc(path, f"Cost Estimate — {project['project_name']}")
        story    = _header(self._settings, s, "CONSTRUCTION COST ESTIMATE REPORT", project)
        story   += [_project_block(project, s)]

        if not estimate:
            story.append(Paragraph("No estimate has been saved for this project yet.", s["RF_Body"]))
            doc.build(story)
            return

        # ── Summary KPI cards as a table ─────────────────────────────────────
        story += _section_bar("Cost Summary at a Glance", s)
        kpis = [
            ("Material Cost",      estimate["material_cost"]),
            ("Labour Cost",        estimate["labour_cost"]),
            ("Equipment Cost",     estimate["equipment_cost"]),
            ("Contractor Margin",  estimate["contractor_margin"]),
            (f"GST ({estimate.get('gst_pct',18):.0f}%)", estimate["gst_amount"]),
        ]
        kpi_data = [[Paragraph(f"<b>{lbl}</b>", s["RF_Body"]), fmt_inr(val)] for lbl, val in kpis]
        kpi_tbl = Table([["Description", "Amount (₹)"]] + kpi_data,
                        colWidths=[USABLE_W - 50*mm, 50*mm])
        kpi_tbl.setStyle(TableStyle([
            ("BACKGROUND",  (0,0), (-1,0), C_STEEL),
            ("TEXTCOLOR",   (0,0), (-1,0), C_WHITE),
            ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",    (0,0), (-1,-1), 9),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white, C_LIGHT]),
            ("BOX",         (0,0), (-1,-1), 0.5, C_NAVY),
            ("INNERGRID",   (0,0), (-1,-1), 0.3, colors.HexColor("#cccccc")),
            ("ALIGN",       (1,0), (1,-1), "RIGHT"),
            ("TOPPADDING",  (0,0), (-1,-1), 6), ("BOTTOMPADDING",(0,0),(-1,-1),6),
            ("LEFTPADDING", (0,0), (-1,-1), 8),
        ]))
        story += [kpi_tbl, Spacer(1,4),
                  _total_row("GRAND TOTAL (incl. GST)", fmt_inr(estimate["grand_total"])),
                  Spacer(1, 10)]

        # ── Charts side by side ───────────────────────────────────────────────
        story += _section_bar("Cost Composition Charts", s)
        labels = ["Material", "Labour", "Equipment", "Margin", "GST"]
        vals   = [estimate["material_cost"], estimate["labour_cost"],
                  estimate["equipment_cost"], estimate["contractor_margin"],
                  estimate["gst_amount"]]
        pie  = _pie_chart(labels, vals, "Cost Distribution", w_mm=82, h_mm=72)
        bar  = _bar_chart(labels, vals, "Cost Component Breakdown (₹)", w_mm=90, h_mm=72)
        story.append(_side_by_side(pie, "Fig 1 — Cost distribution by component",
                                   bar, "Fig 2 — Absolute amounts per component", s))
        story.append(Spacer(1,10))

        # ── Per-sqft analysis ─────────────────────────────────────────────────
        area = project.get("built_up_area", 0) or 1
        psf  = estimate["grand_total"] / area
        story += _section_bar("Rate Analysis", s)
        rate_data = [
            ["Built-up Area",        f"{area:,.0f} sq.ft"],
            ["Cost per sq.ft (total)",f"₹ {psf:,.2f}"],
            ["Material per sq.ft",   f"₹ {estimate['material_cost']/area:,.2f}"],
            ["Labour per sq.ft",     f"₹ {estimate['labour_cost']/area:,.2f}"],
        ]
        rate_tbl = Table([["Parameter","Value"]] + rate_data,
                         colWidths=[USABLE_W - 50*mm, 50*mm])
        rate_tbl.setStyle(TableStyle([
            ("BACKGROUND",  (0,0),(-1,0), C_STEEL),
            ("TEXTCOLOR",   (0,0),(-1,0), C_WHITE),
            ("FONTNAME",    (0,0),(-1,0), "Helvetica-Bold"),
            ("FONTSIZE",    (0,0),(-1,-1), 9),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,C_LIGHT]),
            ("BOX",         (0,0),(-1,-1), 0.5, C_NAVY),
            ("INNERGRID",   (0,0),(-1,-1), 0.3, colors.HexColor("#cccccc")),
            ("ALIGN",       (1,0),(1,-1), "RIGHT"),
            ("TOPPADDING",  (0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6),
            ("LEFTPADDING", (0,0),(-1,-1),8),
        ]))
        story += [rate_tbl, Spacer(1,10)]

        # ── Detailed material table ───────────────────────────────────────────
        story += _section_bar("Detailed Material Quantity Schedule", s)
        items = estimate.get("items", [])
        hdrs = ["#", "Item Description", "Quantity", "Unit", "Rate (₹)", "Amount (₹)"]
        cws  = [8*mm, 72*mm, 22*mm, 16*mm, 28*mm, 30*mm]
        rows = [[str(i+1), it["item_name"], f"{it['quantity']:,.2f}",
                 it["unit"], fmt_inr(it["rate"]), fmt_inr(it["amount"])]
                for i, it in enumerate(items)]
        story += [_std_table(hdrs, rows, cws, right_cols=[4,5]),
                  Spacer(1,4),
                  _highlight_row("Sub-total (Materials)", fmt_inr(estimate["material_cost"])),
                  Spacer(1,12)]

        # ── Footer note ───────────────────────────────────────────────────────
        story.append(HRFlowable(width="100%", thickness=0.5,
                                color=colors.HexColor("#cccccc"), spaceBefore=4))
        story.append(Paragraph(
            "This estimate is based on prevailing Indian market rates and project parameters entered "
            "in ArchForge Pro. Actual costs may vary depending on site conditions, material "
            "availability, and contractor negotiations. This document is intended for planning "
            "purposes only.", s["RF_Note"]
        ))
        doc.build(story)

    # ── 2. BOQ Report ────────────────────────────────────────────────────────
    def generate_boq_report(self, project_id: int, path: str) -> None:
        self.export_boq_pdf(project_id, path)

    def export_boq_pdf(self, project_id: int, path: str) -> None:
        project   = get_project(project_id)
        boq_items = get_boq(project_id)
        s         = _styles()
        doc       = _doc(path, f"BOQ — {project['project_name']}")
        story     = _header(self._settings, s, "BILL OF QUANTITIES (BOQ)", project)
        story    += [_project_block(project, s)]

        if not boq_items:
            story.append(Paragraph("No BOQ items available. Generate a BOQ from the BOQ page first.", s["RF_Body"]))
            doc.build(story)
            return

        total = sum(i.get("amount", 0) for i in boq_items)

        # ── BOQ summary chart ─────────────────────────────────────────────────
        # Group by first word of description (rough category)
        from collections import defaultdict
        cat_totals: dict[str, float] = defaultdict(float)
        for item in boq_items:
            cat = item["description"].split()[0] if item.get("description") else "Other"
            cat_totals[cat] += item.get("amount", 0)
        top_cats = sorted(cat_totals.items(), key=lambda x: x[1], reverse=True)[:7]
        if top_cats:
            story += _section_bar("BOQ Cost Distribution", s)
            pie = _pie_chart([c for c,_ in top_cats], [v for _,v in top_cats],
                             "By Work Category", w_mm=85, h_mm=72)
            bar = _bar_chart([c for c,_ in top_cats], [v for _,v in top_cats],
                             "Amount by Category (₹)", w_mm=88, h_mm=72)
            story.append(_side_by_side(
                pie, "Fig 1 — BOQ cost share by category",
                bar, "Fig 2 — Category-wise amounts", s))
            story.append(Spacer(1,10))

        # ── Full BOQ table ────────────────────────────────────────────────────
        story += _section_bar("Complete Bill of Quantities", s)
        hdrs = ["Item No.", "Description of Work", "Qty", "Unit", "Rate (₹)", "Amount (₹)"]
        cws  = [14*mm, 74*mm, 16*mm, 16*mm, 28*mm, 30*mm]
        rows = [[str(it["item_no"]), it["description"],
                 f"{it['quantity']:,.2f}", it["unit"],
                 fmt_inr(it["rate"]), fmt_inr(it["amount"])]
                for it in boq_items]
        story += [_std_table(hdrs, rows, cws, right_cols=[4,5]),
                  Spacer(1,4),
                  _total_row("TOTAL AMOUNT (₹)", fmt_inr(total)),
                  Spacer(1,12)]

        # ── Per-sqft BOQ rate ─────────────────────────────────────────────────
        area = project.get("built_up_area", 0) or 1
        story += _section_bar("Rate Summary", s)
        summary_tbl = Table([
            ["Total BOQ Value",     fmt_inr(total)],
            ["Built-up Area",       f"{area:,.0f} sq.ft"],
            ["Rate per sq.ft",      f"₹ {total/area:,.2f}"],
        ], colWidths=[USABLE_W - 50*mm, 50*mm])
        summary_tbl.setStyle(TableStyle([
            ("FONTSIZE",(0,0),(-1,-1),9),
            ("ROWBACKGROUNDS",(0,0),(-1,-1),[C_LIGHT, colors.white]),
            ("BOX",(0,0),(-1,-1),0.5,C_NAVY),
            ("INNERGRID",(0,0),(-1,-1),0.3,colors.HexColor("#cccccc")),
            ("FONTNAME",(0,0),(0,-1),"Helvetica-Bold"),
            ("ALIGN",(1,0),(1,-1),"RIGHT"),
            ("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6),
            ("LEFTPADDING",(0,0),(-1,-1),8),
        ]))
        story.append(summary_tbl)
        story.append(Spacer(1,12))
        story.append(HRFlowable(width="100%",thickness=0.5,
                                color=colors.HexColor("#cccccc"),spaceBefore=4))
        story.append(Paragraph(
            "Rates are based on current Indian Standard Schedule of Rates (DSR). "
            "All quantities are approximate and subject to site measurement. "
            "Contractor to verify quantities before tendering.", s["RF_Note"]
        ))
        doc.build(story)

    def export_boq_excel(self, project_id: int, path: str) -> None:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        project   = get_project(project_id)
        boq_items = get_boq(project_id)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BOQ"
        ws["A1"] = self._settings.get("company_name","ArchForge Pro")
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = "BILL OF QUANTITIES"
        ws["A2"].font = Font(bold=True, size=12)
        ws["A3"] = f"Project: {project['project_name']}  |  Client: {project['client_name']}"
        headers = ["#","Description","Quantity","Unit","Rate (₹)","Amount (₹)"]
        hfill   = PatternFill("solid", fgColor="1E3A5F")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = hfill
            cell.alignment = Alignment(horizontal="center")
        total = 0.0
        for ri, item in enumerate(boq_items, 6):
            ws.cell(ri,1,item["item_no"]); ws.cell(ri,2,item["description"])
            ws.cell(ri,3,item["quantity"]); ws.cell(ri,4,item["unit"])
            ws.cell(ri,5,item["rate"]);     ws.cell(ri,6,item["amount"])
            if ri % 2 == 0:
                for c in range(1,7):
                    ws.cell(ri,c).fill = PatternFill("solid",fgColor="F0F4F8")
            total += item.get("amount",0)
        tr = len(boq_items)+6
        ws.cell(tr,5,"TOTAL").font = Font(bold=True)
        ws.cell(tr,6,total).font   = Font(bold=True)
        for col, w in enumerate([6,50,12,10,16,16],1):
            ws.column_dimensions[ws.cell(1,col).column_letter].width = w
        wb.save(path)

    # ── 3. Expense Report ────────────────────────────────────────────────────
    def generate_expense_report(self, project_id: int, path: str) -> None:
        project  = get_project(project_id)
        expenses = get_expenses(project_id)
        summary  = get_expense_summary(project_id)
        monthly  = get_monthly_spending(project_id)
        s        = _styles()
        doc      = _doc(path, f"Expense Report — {project['project_name']}")
        story    = _header(self._settings, s, "EXPENSE REPORT", project)
        story   += [_project_block(project, s)]

        total = summary["total"]
        by_cat = summary.get("by_category", {})

        # ── KPI strip ────────────────────────────────────────────────────────
        story += _section_bar("Expenditure Summary", s)
        kpi_data = [[Paragraph(f"<b>{k}</b>", s["RF_Body"]), fmt_inr(v)]
                    for k, v in by_cat.items()]
        kpi_data.append([Paragraph("<b>Grand Total</b>", s["RF_Body"]),
                          Paragraph(f"<b>{fmt_inr(total)}</b>", s["RF_Body"])])
        kpi_tbl = Table([["Category","Amount (₹)"]] + kpi_data,
                        colWidths=[USABLE_W-50*mm, 50*mm])
        kpi_tbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),C_STEEL),
            ("TEXTCOLOR",(0,0),(-1,0),C_WHITE),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
            ("FONTSIZE",(0,0),(-1,-1),9),
            ("ROWBACKGROUNDS",(0,1),(-1,-2),[colors.white,C_LIGHT]),
            ("BACKGROUND",(0,-1),(-1,-1),C_LIGHT),
            ("BOX",(0,0),(-1,-1),0.5,C_NAVY),
            ("INNERGRID",(0,0),(-1,-1),0.3,colors.HexColor("#cccccc")),
            ("ALIGN",(1,0),(1,-1),"RIGHT"),
            ("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6),
            ("LEFTPADDING",(0,0),(-1,-1),8),
        ]))
        story += [kpi_tbl, Spacer(1,10)]

        # ── Charts ────────────────────────────────────────────────────────────
        if by_cat:
            story += _section_bar("Expenditure Analysis", s)
            pie = _pie_chart(list(by_cat.keys()), list(by_cat.values()),
                             "Spend by Category", w_mm=82, h_mm=72)
            if monthly:
                months = [m["month"] for m in monthly]
                mvals  = [m["total"]  for m in monthly]
                line = _line_chart(months, mvals, "Monthly Spend Trend (₹)",
                                   w_mm=90, h_mm=72)
                story.append(_side_by_side(
                    pie, "Fig 1 — Category-wise expenditure",
                    line,"Fig 2 — Monthly spending trend", s))
            else:
                pie.drawWidth  = USABLE_W
                pie.drawHeight = 80*mm
                story.append(pie)
            story.append(Spacer(1,10))

        # ── Itemised ledger ───────────────────────────────────────────────────
        story += _section_bar("Itemised Expense Ledger", s)
        hdrs = ["Date", "Category", "Description", "Receipt Ref", "Amount (₹)"]
        cws  = [22*mm, 28*mm, 72*mm, 22*mm, 30*mm]
        rows = [[fmt_date(e["expense_date"]), e["category"],
                 e["description"], e.get("receipt_ref","—"), fmt_inr(e["amount"])]
                for e in expenses]
        story += [_std_table(hdrs, rows, cws, right_cols=[4]),
                  Spacer(1,4),
                  _total_row("TOTAL EXPENDITURE (₹)", fmt_inr(total)),
                  Spacer(1,12)]

        story.append(HRFlowable(width="100%",thickness=0.5,
                                color=colors.HexColor("#cccccc"),spaceBefore=4))
        story.append(Paragraph(
            "All expenses are as recorded in ArchForge Pro. Amounts are inclusive of applicable taxes. "
            "Please retain original receipts for audit purposes.", s["RF_Note"]
        ))
        doc.build(story)

    # ── 4. Variance Analysis Report ──────────────────────────────────────────
    def generate_variance_report(self, project_id: int, path: str) -> None:
        project   = get_project(project_id)
        estimate  = get_estimate(project_id)
        summary   = get_expense_summary(project_id)
        s         = _styles()
        doc       = _doc(path, f"Variance Analysis — {project['project_name']}")
        story     = _header(self._settings, s, "COST VARIANCE ANALYSIS REPORT", project)
        story    += [_project_block(project, s)]

        est_total = estimate["grand_total"] if estimate else 0
        actual    = summary["total"]
        variance  = actual - est_total
        pct       = (variance / est_total * 100) if est_total else 0
        over      = variance > 0

        # ── Key metrics ───────────────────────────────────────────────────────
        story += _section_bar("Variance Summary", s)
        status_col = C_RED if over else C_GREEN
        status_txt = "OVER BUDGET" if over else "WITHIN BUDGET"
        rows = [
            ["Estimated Cost (Budget)",   fmt_inr(est_total)],
            ["Actual Expenditure to Date",fmt_inr(actual)],
            ["Absolute Variance",         fmt_inr(abs(variance))],
            ["Variance Percentage",       f"{abs(pct):.2f}%"],
            ["Remaining Budget",          fmt_inr(max(0, est_total - actual))],
            ["Budget Utilisation",        f"{min(100, actual/est_total*100) if est_total else 0:.1f}%"],
        ]
        var_tbl = Table([["Metric","Value"]] + rows,
                        colWidths=[USABLE_W-50*mm, 50*mm])
        style_cmds = [
            ("BACKGROUND",(0,0),(-1,0),C_STEEL),
            ("TEXTCOLOR",(0,0),(-1,0),C_WHITE),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
            ("FONTSIZE",(0,0),(-1,-1),9),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,C_LIGHT]),
            ("BOX",(0,0),(-1,-1),0.5,C_NAVY),
            ("INNERGRID",(0,0),(-1,-1),0.3,colors.HexColor("#cccccc")),
            ("ALIGN",(1,0),(1,-1),"RIGHT"),
            ("FONTNAME",(0,1),(0,-1),"Helvetica-Bold"),
            ("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6),
            ("LEFTPADDING",(0,0),(-1,-1),8),
        ]
        var_tbl.setStyle(TableStyle(style_cmds))
        story += [var_tbl, Spacer(1,4),
                  _highlight_row(f"STATUS: {status_txt}",
                                 f"{'▲' if over else '▼'} {abs(pct):.1f}% vs budget",
                                 status_col),
                  Spacer(1,10)]

        # ── Visual comparison ─────────────────────────────────────────────────
        story += _section_bar("Visual Comparison", s)
        if est_total > 0 or actual > 0:
            comp_bar = _bar_chart(
                ["Estimated Budget", "Actual Spend", "Remaining"],
                [est_total, actual, max(0, est_total - actual)],
                "Budget vs Actual (₹)",
                color_list=["#2c4a6e","#c8620a","#2e7d4f"],
                w_mm=88, h_mm=72,
            )
            # Budget utilisation donut
            util_pct = min(100, actual/est_total*100) if est_total else 0
            fig, ax = plt.subplots(figsize=(82/25.4, 72/25.4))
            wedge_vals = [util_pct, max(0, 100 - util_pct)]
            wedge_cols = ["#c8620a" if over else "#2e7d4f", "#e8ecf0"]
            ax.pie(wedge_vals, colors=wedge_cols, startangle=90,
                   wedgeprops=dict(width=0.42, edgecolor="white", linewidth=1.5))
            ax.text(0, 0, f"{util_pct:.1f}%", ha="center", va="center",
                    fontsize=16, fontweight="bold", color=wedge_cols[0])
            ax.text(0, -0.55, "Budget Used", ha="center", va="center",
                    fontsize=8, color="#555555")
            ax.set_title("Budget Utilisation", fontsize=9, fontweight="bold",
                         color="#1e3a5f", pad=8)
            util_img = _chart_to_image(fig, 82, 72)
            story.append(_side_by_side(
                util_img, "Fig 1 — Budget utilisation",
                comp_bar, "Fig 2 — Estimate vs actual spend", s))
            story.append(Spacer(1,10))

        # ── Component-level breakdown ─────────────────────────────────────────
        if estimate:
            story += _section_bar("Estimated vs Actual — Component Breakdown", s)
            by_cat = summary.get("by_category", {})
            comp_rows = [
                ["Material",     estimate["material_cost"],  by_cat.get("Material",0)],
                ["Labour",       estimate["labour_cost"],    by_cat.get("Labour",0)],
                ["Equipment",    estimate["equipment_cost"], by_cat.get("Equipment",0)],
                ["Contractor",   estimate["contractor_margin"], by_cat.get("Contractor",0)],
            ]
            hdrs = ["Component","Estimated (₹)","Actual (₹)","Variance (₹)","Status"]
            tbl_rows = []
            for comp, est_v, act_v in comp_rows:
                var_v = act_v - est_v
                st = "Over" if var_v > 0 else ("On Track" if act_v > 0 else "—")
                tbl_rows.append([comp, fmt_inr(est_v), fmt_inr(act_v),
                                  fmt_inr(abs(var_v)), st])
            comp_tbl = _std_table(hdrs, tbl_rows,
                                  [35*mm,38*mm,38*mm,38*mm,29*mm],
                                  right_cols=[1,2,3])
            story += [comp_tbl, Spacer(1,12)]

        story.append(HRFlowable(width="100%",thickness=0.5,
                                color=colors.HexColor("#cccccc"),spaceBefore=4))
        story.append(Paragraph(
            "Variance analysis is based on expenses recorded in ArchForge Pro against the saved "
            "cost estimate. For accurate tracking, ensure all expenses are logged promptly. "
            "Positive variance indicates overspend; negative indicates savings.", s["RF_Note"]
        ))
        doc.build(story)

    # ── 5. Project Summary Report ────────────────────────────────────────────
    def generate_project_summary_report(self, project_id: int, path: str) -> None:
        project     = get_project(project_id)
        estimate    = get_estimate(project_id)
        expense_sum = get_expense_summary(project_id)
        phases      = get_phases(project_id)
        s           = _styles()
        doc         = _doc(path, f"Project Summary — {project['project_name']}")
        story       = _header(self._settings, s, "PROJECT SUMMARY REPORT", project)
        story      += [_project_block(project, s)]

        # ── Financial overview ────────────────────────────────────────────────
        est_total = estimate["grand_total"] if estimate else 0
        actual    = expense_sum["total"]
        remaining = max(0, est_total - actual)
        util_pct  = (actual / est_total * 100) if est_total else 0

        story += _section_bar("Financial Overview", s)
        fin_rows = [
            ["Estimated Budget",          fmt_inr(est_total)],
            ["Actual Expenditure",        fmt_inr(actual)],
            ["Budget Remaining",          fmt_inr(remaining)],
            ["Budget Utilisation",        f"{util_pct:.1f}%"],
            ["Cost per sq.ft (estimate)", f"₹ {est_total/max(1,project.get('built_up_area',1)):,.2f}"],
        ]
        fin_tbl = Table([["Metric","Value"]] + fin_rows,
                        colWidths=[USABLE_W-50*mm, 50*mm])
        fin_tbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),C_STEEL),
            ("TEXTCOLOR",(0,0),(-1,0),C_WHITE),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
            ("FONTSIZE",(0,0),(-1,-1),9),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,C_LIGHT]),
            ("BOX",(0,0),(-1,-1),0.5,C_NAVY),
            ("INNERGRID",(0,0),(-1,-1),0.3,colors.HexColor("#cccccc")),
            ("FONTNAME",(0,1),(0,-1),"Helvetica-Bold"),
            ("ALIGN",(1,0),(1,-1),"RIGHT"),
            ("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6),
            ("LEFTPADDING",(0,0),(-1,-1),8),
        ]))
        story += [fin_tbl, Spacer(1,10)]

        # ── Cost composition charts ───────────────────────────────────────────
        if estimate:
            story += _section_bar("Cost Composition", s)
            labels = ["Material","Labour","Equipment","Margin","GST"]
            vals   = [estimate["material_cost"], estimate["labour_cost"],
                      estimate["equipment_cost"], estimate["contractor_margin"],
                      estimate["gst_amount"]]
            pie  = _pie_chart(labels, vals, "Estimated Cost Breakdown", w_mm=82, h_mm=72)
            bar2 = _bar_chart(
                ["Budget","Actual","Remaining"],
                [est_total, actual, remaining],
                "Budget Tracker (₹)",
                color_list=["#2c4a6e","#c8620a","#2e7d4f"],
                w_mm=90, h_mm=72,
            )
            story.append(_side_by_side(
                pie, "Fig 1 — Estimated cost breakdown",
                bar2,"Fig 2 — Budget vs actual", s))
            story.append(Spacer(1,10))

        # ── Phase progress table + Gantt ──────────────────────────────────────
        if phases:
            story += _section_bar("Construction Phase Progress", s)
            ph_hdrs = ["Phase","Planned Start","Planned End","Actual Start","Actual End","Completion","Status"]
            ph_cws  = [38*mm,22*mm,22*mm,22*mm,22*mm,18*mm,24*mm]
            ph_rows = []
            for p in phases:
                ph_rows.append([
                    p["phase_name"],
                    fmt_date(p.get("planned_start")),
                    fmt_date(p.get("planned_end")),
                    fmt_date(p.get("actual_start")) or "—",
                    fmt_date(p.get("actual_end"))   or "—",
                    f"{p.get('completion_pct',0) or 0:.0f}%",
                    p.get("status","Not Started"),
                ])
            story += [_std_table(ph_hdrs, ph_rows, ph_cws), Spacer(1,8)]

            gantt = _gantt_chart(phases, w_mm=170, h_mm=88)
            story.append(gantt)
            story.append(Paragraph("Fig 3 — Phase Gantt chart (planned vs actual dates)", s["RF_Caption"]))
            story.append(Spacer(1,10))

        # ── Expense by category ───────────────────────────────────────────────
        by_cat = expense_sum.get("by_category", {})
        if by_cat:
            story += _section_bar("Actual Expenditure by Category", s)
            cat_rows = [[cat, fmt_inr(val)] for cat, val in by_cat.items()]
            cat_tbl = Table([["Category","Amount (₹)"]] + cat_rows,
                            colWidths=[USABLE_W-50*mm, 50*mm])
            cat_tbl.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,0),C_STEEL),
                ("TEXTCOLOR",(0,0),(-1,0),C_WHITE),
                ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
                ("FONTSIZE",(0,0),(-1,-1),9),
                ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,C_LIGHT]),
                ("BOX",(0,0),(-1,-1),0.5,C_NAVY),
                ("INNERGRID",(0,0),(-1,-1),0.3,colors.HexColor("#cccccc")),
                ("ALIGN",(1,0),(1,-1),"RIGHT"),
                ("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6),
                ("LEFTPADDING",(0,0),(-1,-1),8),
            ]))
            story += [cat_tbl,
                      Spacer(1,4),
                      _total_row("TOTAL EXPENDITURE (₹)", fmt_inr(actual)),
                      Spacer(1,12)]

        story.append(HRFlowable(width="100%",thickness=0.5,
                                color=colors.HexColor("#cccccc"),spaceBefore=4))
        story.append(Paragraph(
            f"Report generated by ArchForge Pro on {datetime.now().strftime('%d %B %Y')}. "
            "All financial data is sourced from the project database. Phase dates and completion "
            "percentages are as last updated on the Timeline page.", s["RF_Note"]
        ))
        doc.build(story)
